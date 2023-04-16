import torch
from DatasetLoader import VideoDataset
import torch.utils.data as Data
from models.slowfast import SlowFast
from models.c3d import C3D
from models.r3d import r3d_18, r3d_34
from models.r2plus1d import r2plus1d_18, r2plus1d_34
from models.resnet3d import r3d_50, r2plus1d_50
import openpyxl
import os


# hyper parameters
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
EPOCH = 100
LR = 0.001
BATCH_SIZE = 32
data_dir = 'datasets'
save_dir = 'results'
if not os.path.exists(save_dir):
    os.mkdir(save_dir)


# data
traindata = VideoDataset(txt='train.txt', filepath=data_dir, n_frame=16)
testdata = VideoDataset(txt='val.txt', filepath=data_dir, n_frame=16)
trainloader = Data.DataLoader(dataset=traindata, batch_size=BATCH_SIZE, shuffle=True, num_workers=1)
testloader = Data.DataLoader(dataset=testdata, batch_size=BATCH_SIZE, shuffle=False, num_workers=1)

# model
model_name = 'slowfast_r50'
dataset = 'ava'
net = SlowFast().to(DEVICE)
loss_func = torch.nn.CrossEntropyLoss()  # loss function
optimizer = torch.optim.SGD(net.parameters(), LR, weight_decay=0.02)  # optimize


def training():
    net.train()
    losses, acc = 0, 0
    total = 0
    for step, (inputs, labels) in enumerate(trainloader):
        inputs, labels = inputs.to(DEVICE), labels.to(DEVICE)
        outputs = net(inputs)
        loss = loss_func(outputs, labels).to(DEVICE)
        losses += loss
        predict = torch.max(outputs.data, 1)[1]
        acc += (predict == labels).sum()
        total += labels.size(0)

        optimizer.zero_grad()  # gradient to zero
        loss.backward()  # backward
        optimizer.step()  # renew weight

        if step == 10:
            break

    losses = losses.to("cpu")
    losses = losses.data.numpy().item() / total
    acc = acc.to("cpu")
    acc = acc.numpy().item() / total
    return [losses, acc]


def testing():
    net.eval()
    loss, acc = 0, 0
    total = 0
    tp, fp, tn, fn = 0, 0, 0, 0
    with torch.no_grad():
        for step, (test_inputs, test_labels) in enumerate(testloader):
            test_inputs, test_labels = test_inputs.to(DEVICE), test_labels.to(DEVICE)
            test_outputs = net(test_inputs)
            loss += loss_func(test_outputs, test_labels).to(DEVICE)
            predict = torch.max(test_outputs.data, 1)[1]
            acc += (predict == test_labels).sum()
            total += test_labels.size(0)

            one = torch.ones(test_labels.size()).to(DEVICE)
            zero = torch.zeros(test_labels.size()).to(DEVICE)
            tp += ((predict == one) & (test_labels == one)).sum()
            fp += ((predict == one) & (test_labels == zero)).sum()
            fn += ((predict == zero) & (test_labels == one)).sum()
            tn += ((predict == zero) & (test_labels == zero)).sum()

            if step == 10:
                break

        loss = loss.to("cpu")
        loss = loss.data.numpy().item() / total
        acc = acc.to("cpu")
        acc = acc.numpy().item() / total
        tp, fp, tn, fn = tp.to("cpu"), fp.to("cpu"), tn.to("cpu"), fn.to("cpu")
        tp, fp, tn, fn = tp.numpy().item(), fp.numpy().item(), tn.numpy().item(), fn.numpy().item()
        return [loss, acc, [tp, fp, tn, fn]]


if __name__ == "__main__":
    train_loss = []
    train_acc = []
    test_loss = []
    test_acc = []
    matrixes = []
    precisions = []
    recalls = []

    # run
    for i in range(EPOCH):
        if (EPOCH+1) % 10 == 0:
            LR = LR / 10
        trainloss, trainacc = training()
        train_loss.append(trainloss), train_acc.append(trainacc)
        testloss, testacc, matrix = testing()
        test_loss.append(testloss), test_acc.append(testacc)
        matrixes.append(matrix)
        print('epoch: {}, trainloss: {:.4f}, trainacc: {:.4f}, testloss: {:.4f}, testacc: {:.4f}'
              .format(i + 1, trainloss, trainacc, testloss, testacc))
        tp, fp, tn, fn = matrix
        precision = 0 if tp == 0 else tp / (tp + fp)
        recall = 0 if tp == 0 else tp / (tp + fn)
        precisions.append(precision)
        recalls.append(recall)
        print("epoch:", i+1, "tp:", tp, "fp:", fp, "tn:", tn, "fn:", fn, "precision: %0.4f" % precision,
              "recall: %0.4f" % recall)

    # save data
    file_name = os.path.join('result', model_name+'.xlsx')
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()
    sheets = workbook.get_sheet_names()
    if dataset in sheets:
        sheet = workbook.get_sheet_by_name(dataset)
    else:
        sheet = workbook.create_sheet(dataset)
    for i in range(1, EPOCH+1):
        sheet.cell(row=i, column=1).value = train_loss[i-1]
        sheet.cell(row=i, column=2).value = train_acc[i-1]
        sheet.cell(row=i, column=3).value = test_loss[i-1]
        sheet.cell(row=i, column=4).value = test_acc[i-1]
        for j in range(1, len(matrixes[0])+1):
            sheet.cell(row=i, column=j+4).value = matrixes[i-1][j-1]
        sheet.cell(row=i, column=9).value = precisions[i-1]
        sheet.cell(row=i, column=10).value = recalls[i-1]
    workbook.save(file_name)
